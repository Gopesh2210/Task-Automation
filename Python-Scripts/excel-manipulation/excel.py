'''
    This script is used for manipulation of excel data
'''

import openpyxl


def readExcel():
    # gets the excel file
    workbook = openpyxl.load_workbook('example.xlsx')

    # gets the sheet
    sheet_names = workbook.get_sheet_names()
    print("Sheets available by name :",sheet_names)

    # get specific sheet & the cell object
    sheet = workbook.get_sheet_by_name('Sheet1')
    cell_value = (sheet['B1'].value)
    print("Cell Data : ",cell_value)

    # acquiring all values of column 2
    for i in range(1,8):
        print("Value of row " + str(i) + "\t" + sheet.cell(row=i,column=2).value)



def editExcel():
    # makes the excel file
    wb = openpyxl.Workbook()

    print("\nYour sheet is created!")
    print("Welcome to the sheet to add your favourtie fruits name\n")


    # gets the sheet
    # sheet_names = wb.get_sheet_names()
    # print("Sheets available by name :",sheet_names)
    # print("\n")
    sheet = wb.get_sheet_by_name('Sheet')
    print("\n")

    # changing sheet name
    title = input("Please insert your sheet name : ")
    sheet.title = title

    # adding values to the excel
    print("Add your 5 favorite fruit names : ")
    fruits = []
    for i in range(0,5):
        fruits.append(str(input()))

    for i in range(0,len(fruits)):
        sheet.cell(row=i+1,column=1).value = i+1
        sheet.cell(row=i+1,column=2).value = fruits[i]

    # saving excel file
    print("\nExcel Created with the name edit-sample.xlsx")
    wb.save('edit-sample.xlsx')

    




# Calling the edit and read excel methods
print("Choose Operation : \n")
choice = input("1: Read Excel\n2: Edit Excel\n\n")

if(choice == '1'):
    readExcel()
else:
    editExcel()

